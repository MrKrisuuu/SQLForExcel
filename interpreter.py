import AST
from visit import *
import openpyxl

ROW_COLUMN_NAMES = 1


def insert_first_row(sheet, row, rownum):
    sheet.cell(row=rownum, column=1).value = "id"
    for i in range(len(row)):
        sheet.cell(row=rownum, column=i+2).value = row[i]


def insert_row(sheet, row, rownum):
    if rownum == 2:
        id = 1
    else:
        id = sheet.cell(row=rownum-1, column=1).value +1
    row = [id] + row
    for i in range(len(row)):
        sheet.cell(row=rownum, column=i+1).value = row[i]


def get_columns_names(sheet, shortname):  # {column_name -> id_column}
    row = sheet[ROW_COLUMN_NAMES]
    data = {}
    for (i, cell) in enumerate(row):
        data[shortname + "." + cell.value] = i
    return data


def read_row(sheet, column_names, rownum):  # {column_name -> value}
    row = sheet[rownum]
    modified_row = {}
    for column_name in column_names.keys():
        modified_row[column_name] = row[column_names[column_name]].value
    return modified_row


class Interpreter(object):
    def __init__(self, name):
        self.name = name
        self.current_row = None

    # BASIC STUFF

    @on('node')
    def visit(self, node):
        pass

    @when(AST.Instructions)
    def visit(self, node):
        self.visit(node.left)
        self.visit(node.right)

    # CRUD

    @when(AST.CreateTable)
    def visit(self, node):
        (fullname, shortname) = self.visit(node.table)
        columns = self.visit(node.columns)
        database = openpyxl.load_workbook(self.name)
        if fullname in database.get_sheet_names():
            table = database.get_sheet_by_name(fullname)
            database.remove_sheet(table)
        table = database.create_sheet(fullname)
        insert_first_row(table, columns, 1)
        database.save(self.name)

    @when(AST.InsertInto)
    def visit(self, node):
        (fullname, shortname) = self.visit(node.table)
        values = self.visit(node.variables)
        database = openpyxl.load_workbook(self.name)
        table = database.get_sheet_by_name(fullname)
        for value in values:
            insert_row(table, value, table.max_row + 1)
        database.save(self.name)

    @when(AST.SelectFrom)
    def visit(self, node):
        # SELECT
        columns_in_select = self.visit(node.columns)  # [column_name]

        # FROM
        (fullname, shortname) = self.visit(node.table)  # (fullname, shortname)
        database = openpyxl.load_workbook(self.name)
        table = database.get_sheet_by_name(fullname)

        column_names = get_columns_names(table, shortname)  # {column_name -> id_column}

        current_result = []   # [{column_name -> value}]
        for rownum in range(2, table.max_row + 1):
            row = read_row(table, column_names, rownum)
            current_result.append(row)

        # JOIN ON
        left = current_result
        joinings = self.visit(node.joining)
        if joinings is not None:
            for (join_type, (fullname, shortname), condition_node) in joinings:
                table = database.get_sheet_by_name(fullname)
                column_names = get_columns_names(table, shortname)
                right = []
                for rownum in range(2, table.max_row + 1):
                    row = read_row(table, column_names, rownum)
                    right.append(row)
                tmp_result = []
                if join_type is None:
                    for left_row in left:
                        for right_row in right:
                            joined_row = left_row | right_row
                            self.current_row = joined_row
                            condition = self.visit(condition_node)
                            self.current_row = None
                            if condition:
                                tmp_result.append(joined_row)
                    left = tmp_result
                if join_type == "LEFT":
                    for left_row in left:
                        added = 0
                        for right_row in right:
                            joined_row = left_row | right_row
                            self.current_row = joined_row
                            condition = self.visit(condition_node)
                            self.current_row = None
                            if condition:
                                added = 1
                                tmp_result.append(joined_row)
                        if added == 0:
                            joined_row = left_row | dict.fromkeys(right[0], None)
                            tmp_result.append(joined_row)
                    left = tmp_result
                if join_type == "RIGHT":
                    raise Exception("RIGHT is not implemented.")
                if join_type == "FULL":
                    raise Exception("FULL is not implemented.")
        current_result = left

        # WHERE
        tmp_result = []
        for row in current_result:
            self.current_row = row
            condition = self.visit(node.whering)
            self.current_row = None
            if condition is None:
                tmp_result.append(row)
            elif condition:
                tmp_result.append(row)
        current_result = tmp_result

        # GROUP BY
        columns_in_group_by = self.visit(node.grouping)
        if columns_in_group_by is not None:
            tmp_result = {}
            for row in current_result:
                lis = []
                for column in columns_in_group_by:
                    lis.append(row[column])
                values_group_by = tuple(lis)
                if values_group_by in tmp_result.keys():
                    tmp_values = tmp_result[values_group_by]
                    for current_column in columns_in_select:
                        if current_column not in columns_in_group_by:
                            (function, col) = current_column
                            if col == "*" or row[col] is not None:
                                match function:
                                    case "COUNT":
                                        tmp_values[current_column] += 1
                                    case "SUM":
                                        tmp_values[current_column] += row[col]
                                    case "AVG":
                                        tmp_values[current_column] = (tmp_values[current_column][0] + row[col], tmp_values[current_column][1] + 1)
                                    case "MIN":
                                        if tmp_values[current_column] is None:
                                            tmp_values[current_column] = row[col]
                                        else:
                                            tmp_values[current_column] = min(tmp_values[current_column], row[col])
                                    case "MAX":
                                        if tmp_values[current_column] is None:
                                            tmp_values[current_column] = row[col]
                                        else:
                                            tmp_values[current_column] = max(tmp_values[current_column], row[col])
                else:
                    tmp_values = {}
                    for current_column in columns_in_select:
                        if current_column not in columns_in_group_by:
                            (function, col) = current_column
                            if col == "*" or row[col] is not None:
                                match function:
                                    case "COUNT":
                                        tmp_values[current_column] = 1
                                    case "SUM":
                                        tmp_values[current_column] = row[col]
                                    case "AVG":
                                        tmp_values[current_column] = (row[col], 1)
                                    case "MIN":
                                        tmp_values[current_column] = row[col]
                                    case "MAX":
                                        tmp_values[current_column] = row[col]
                                    case _:
                                        raise Exception("Wrong type of agregate function.")
                            else:
                                match function:
                                    case "COUNT":
                                        tmp_values[current_column] = 0
                                    case "SUM":
                                        tmp_values[current_column] = 0
                                    case "AVG":
                                        tmp_values[current_column] = (0, 0)
                                    case "MIN":
                                        tmp_values[current_column] = None
                                    case "MAX":
                                        tmp_values[current_column] = None
                                    case _:
                                        raise Exception("Wrong type of agregate function.")
                    tmp_result[values_group_by] = tmp_values
            for values_group_by in tmp_result.keys():
                tmp_values = tmp_result[values_group_by]
                for current_column in columns_in_select:
                    if current_column not in columns_in_group_by:
                        (function, col) = current_column
                        if function == "AVG":
                            if tmp_values[current_column][1] == 0:
                                tmp_values[current_column] = None
                            else:
                                tmp_values[current_column] = tmp_values[current_column][0] / tmp_values[current_column][1]
                tmp_result[values_group_by] = tmp_values
            tmp_result_2 = []
            for row in tmp_result.keys():
                new_dict = {}
                for (colnum, column) in enumerate(columns_in_group_by):
                    new_dict[column] = row[colnum]
                new_dict = new_dict | tmp_result[row]
                tmp_result_2.append(new_dict)
            current_result = tmp_result_2

        # HAVING
        tmp_result = []
        for row in current_result:
            self.current_row = row
            condition = self.visit(node.having)
            self.current_row = None
            if condition is None:
                tmp_result.append(row)
            elif condition:
                tmp_result.append(row)
        current_result = tmp_result

        # ORDER BY
        ordering = self.visit(node.ordering)
        if ordering is not None:
            for column in reversed(ordering):
                current_result.sort(key=lambda x: x[column])

        # RETURN
        # print(columns_in_select)
        final_result = []
        for row in current_result:
            line = []
            for column in columns_in_select:
                if column == "*":
                    line += list(row.values())
                else:
                    line.append(row[column])
            # print(line)
            final_result.append(line)
        # print()
        return final_result

    @when(AST.Update)
    def visit(self, node):
        (fullname, shortname) = self.visit(node.table)
        sets = self.visit(node.sets)
        database = openpyxl.load_workbook(self.name)
        table = database.get_sheet_by_name(fullname)
        column_names = get_columns_names(table, shortname)
        for rownum in range(2, table.max_row + 1):
            row = read_row(table, column_names, rownum)
            self.current_row = row
            condition = self.visit(node.condition)
            self.current_row = None
            if condition:
                for (column_name, value) in sets:
                    i = column_names[column_name]
                    table.cell(row=rownum, column=i + 1).value = value
        database.save(self.name)

    @when(AST.DeleteFrom)
    def visit(self, node):
        (fullname, shortname) = self.visit(node.table)
        database = openpyxl.load_workbook(self.name)
        table = database.get_sheet_by_name(fullname)
        column_names = get_columns_names(table, shortname)
        dels = 0
        for rownum in range(2, table.max_row + 1):
            row = read_row(table, column_names, rownum - dels)
            self.current_row = row
            condition = self.visit(node.condition)
            self.current_row = None
            if condition:
                table.delete_rows(rownum - dels, 1)
                dels += 1
        database.save(self.name)

    # SELECT

    @when(AST.Aggregates)
    def visit(self, node):
        left = self.visit(node.left)
        right = self.visit(node.right)
        if right is None:
            return [left]
        return [left] + right

    @when(AST.Aggregate)
    def visit(self, node):
        function = self.visit(node.function)
        column = self.visit(node.column)
        return (function, column)

    @when(AST.Table)
    def visit(self, node):
        fullname = self.visit(node.fullname)
        shortname = self.visit(node.shortname)
        return (fullname, shortname)

    @when(AST.JoinsOn)
    def visit(self, node):
        left = self.visit(node.left)
        right = self.visit(node.right)
        if right is None:
            return [left]
        return [left] + right

    @when(AST.JoinOn)
    def visit(self, node):
        join_type = self.visit(node.join_type)
        table = self.visit(node.table)
        return (join_type, table, node.condition)

    @when(AST.Where)
    def visit(self, node):
        condition = self.visit(node.condition)
        return condition

    @when(AST.GroupBy)
    def visit(self, node):
        columns = self.visit(node.columns)
        return columns

    @when(AST.Having)
    def visit(self, node):
        condition = self.visit(node.condition)
        return condition

    @when(AST.OrderBy)
    def visit(self, node):
        columns = self.visit(node.columns)
        return columns

    # DOUBLERS

    # VALUES
    @when(AST.Values)
    def visit(self, node):
        left = self.visit(node.left)
        right = self.visit(node.right)
        if right is None:
            return [left]
        return [left] + right

    @when(AST.Value)
    def visit(self, node):
        value = self.visit(node.value)
        return value

    # VARIABLES
    @when(AST.Variables)
    def visit(self, node):
        left = self.visit(node.left)
        right = self.visit(node.right)
        if right is None:
            return [left]
        return [left] + right

    @when(AST.Variable)
    def visit(self, node):
        variable = self.visit(node.variable)
        return variable

    # SETS
    @when(AST.Sets)
    def visit(self, node):
        left = self.visit(node.left)
        right = self.visit(node.right)
        if right is None:
            return [left]
        return [left] + right

    @when(AST.Set)
    def visit(self, node):
        column = self.visit(node.column)
        variable = self.visit(node.variable)
        return (column, variable)

    # COLUMNS
    @when(AST.Columns)
    def visit(self, node):
        left = self.visit(node.left)
        right = self.visit(node.right)
        if right is None:
            return [left]
        return [left] + right

    @when(AST.Column)
    def visit(self, node):
        column = self.visit(node.column)
        return column

    # OTHERS

    @when(AST.Condition)
    def visit(self, node):
        (left, left_type) = self.visit(node.left)
        op = node.op
        (right, right_type) = self.visit(node.right)

        if left is None or right is None:
            return False

        left_value = None
        right_value = None

        if left_type == "variable":
            left_value = left
        elif left_type == "column":
            left_value = self.current_row[left]
        elif left_type == "aggregate":
            left_value = self.current_row[left]
        else:
            raise Exception(f"Wrong type: {left_type}")

        if right_type == "variable":
            right_value = right
        elif right_type == "column":
            right_value = self.current_row[right]
        elif right_type == "aggregate":
            right_value = self.current_row[right]
        else:
            raise Exception(f"Wrong type: {right_type}")

        if op == "==":
            return left_value == right_value
        if op == "!=":
            return left_value != right_value
        if op == "<":
            return left_value < right_value
        if op == ">":
            return left_value > right_value
        if op == "<=":
            return left_value <= right_value
        if op == ">=":
            return left_value >= right_value
        else:
            raise Exception(f"Wrong operator: {op}")

    @when(AST.Side)
    def visit(self, node):
        side = self.visit(node.side)
        return (side, node.type)

    @when(AST.NONE)
    def visit(self, node):
        return None

    @when(AST.ID)
    def visit(self, node):
        return node.name

    @when(AST.STRING)
    def visit(self, node):
        return str(node.value)

    @when(AST.NUMBER)
    def visit(self, node):
        return float(node.value)
